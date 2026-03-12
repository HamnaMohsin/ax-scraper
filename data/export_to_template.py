"""
export_to_template.py
Reads categorized products from SQLite and writes per-category .xlsm files
using the Octopia template.

Usage (CLI):
    python3 export_to_template.py [--only-new]
"""

import os
import re
import sys
import shutil
import argparse
import sqlite3
from datetime import datetime

from openpyxl import load_workbook

# ── Paths ─────────────────────────────────────────────────────────────────────

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
DB_PATH      = os.path.join(BASE_DIR, "products.db")
TEMPLATE_PATH = os.path.join(BASE_DIR, "pdt_template_fr-FR_20260305_090255.xlsm")
OUTPUT_DIR   = os.path.join(BASE_DIR, "output_templates")

# ── Constants ─────────────────────────────────────────────────────────────────

TITLE_MAX         = 132
DESCRIPTION_MAX   = 2000
DESC_MARKETING_MAX = 5000
REF_MAX           = 50
IMAGES_MAX        = 6
FIRST_DATA_ROW    = 11   # row where product data starts in template

# Column numbers (1-indexed) confirmed from template rows 4–5
COLUMN_MAP = {
    "sellerProductReference":  2,
    "title":                   3,
    "description":             4,
    "image_1":                 5,
    "richMarketingDescription": 9,   # confirmed: row 4 col 9
    "image_2":                10,
    "image_3":                11,
    "image_4":                12,
    "image_5":                13,
    "image_6":                14,
}

# ── Filename helpers ──────────────────────────────────────────────────────────

def get_leaf_category(category_name: str) -> str:
    """
    Extract the last segment from a slash-separated category path.
    'FURNITURE/OFFICE FURNITURE/DESK - DESK RISER' → 'DESK - DESK RISER'
    """
    if not category_name:
        return ""
    return category_name.strip().split("/")[-1].strip()


def make_safe_filename(category_id: str, category_name: str) -> str:
    """
    Build sanitized filename: '{code} - {leaf}'.
    e.g. '0G0901 - ELECTRONIC CAMERA MODULE'
    Characters illegal in filenames are removed.
    """
    leaf = get_leaf_category(category_name)
    raw  = f"{category_id} - {leaf}" if leaf else category_id
    for ch in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
        raw = raw.replace(ch, '')
    return raw.strip()


# ── Image URL cleaning ────────────────────────────────────────────────────────

def strip_query(url: str) -> str:
    """Remove query string from image URL (Octopia rejects URLs with ?params)."""
    if not url:
        return ""
    return url.split("?")[0]


# ── Database query ────────────────────────────────────────────────────────────

def load_products(db_path: str = DB_PATH, only_new: bool = False) -> dict:
    """
    Returns a dict keyed by category_id, each value being:
      {
        "category_name": str,
        "products": [ {product fields...}, ... ]
      }

    If only_new=True, only products with exported_at IS NULL are included.
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    base_sql = """
        SELECT
            pf.product_id,
            pf.title            AS original_title,
            pf.description      AS original_description,
            pf.images,
            pr.enhanced_title,
            pr.enhanced_description,
            pr.description_marketing,
            ca.category_id,
            ca.assigned_category
        FROM product_fetched pf
        LEFT JOIN product_refined     pr ON pr.product_id = pf.product_id
        LEFT JOIN category_assignment ca ON ca.product_id = pf.product_id
        WHERE ca.category_id IS NOT NULL
    """

    if only_new:
        base_sql += " AND pf.exported_at IS NULL"

    base_sql += " ORDER BY ca.category_id, pf.product_id"

    cur.execute(base_sql)
    rows = cur.fetchall()
    conn.close()

    categories: dict = {}
    for row in rows:
        cat_id   = row["category_id"]
        cat_name = row["assigned_category"] or ""

        if cat_id not in categories:
            categories[cat_id] = {"category_name": cat_name, "products": []}

        import json as _json
        images_raw = row["images"]
        if isinstance(images_raw, str):
            try:
                images = _json.loads(images_raw)
            except Exception:
                images = []
        elif isinstance(images_raw, list):
            images = images_raw
        else:
            images = []

        categories[cat_id]["products"].append({
            "product_id":             row["product_id"],
            "original_title":         row["original_title"] or "",
            "original_description":   row["original_description"] or "",
            "enhanced_title":         row["enhanced_title"] or "",
            "enhanced_description":   row["enhanced_description"] or "",
            "description_marketing":  row["description_marketing"] or "",
            "images":                 images,
        })

    return categories


# ── Template writer ───────────────────────────────────────────────────────────

def write_category_file(
    category_id:   str,
    category_name: str,
    products:      list,
    out_path:      str,
    template_path: str = TEMPLATE_PATH,
    append_mode:   bool = False,
) -> list:
    """
    Write products into a .xlsm copy of the Octopia template.

    append_mode=True: open existing file and find last written row.
    append_mode=False: copy template fresh and start from FIRST_DATA_ROW.

    Returns list of product_ids written (as int).
    """
    if append_mode and os.path.exists(out_path):
        wb = load_workbook(out_path, keep_vba=True)
    else:
        shutil.copy2(template_path, out_path)
        wb = load_workbook(out_path, keep_vba=True)

    ws = wb.active

    # Find the first empty data row
    if append_mode:
        excel_row = FIRST_DATA_ROW
        while ws.cell(row=excel_row, column=COLUMN_MAP["sellerProductReference"]).value not in (None, ""):
            excel_row += 1
    else:
        excel_row = FIRST_DATA_ROW

    written_ids = []

    for product in products:
        pid   = product["product_id"]
        title = (product["enhanced_title"] or product["original_title"])[:TITLE_MAX]
        desc  = (product["enhanced_description"] or product["original_description"])[:DESCRIPTION_MAX]
        desc_marketing = str(product.get("description_marketing") or "")[:DESC_MARKETING_MAX]

        images = [strip_query(u) for u in product["images"] if u]
        images = [u for u in images if u][:IMAGES_MAX]

        # Write cells
        ws.cell(row=excel_row, column=COLUMN_MAP["sellerProductReference"]).value = str(pid)[:REF_MAX]
        ws.cell(row=excel_row, column=COLUMN_MAP["title"]).value                  = title
        ws.cell(row=excel_row, column=COLUMN_MAP["description"]).value            = desc
        ws.cell(row=excel_row, column=COLUMN_MAP["richMarketingDescription"]).value = desc_marketing

        img_keys = ["image_1", "image_2", "image_3", "image_4", "image_5", "image_6"]
        for i, key in enumerate(img_keys):
            ws.cell(row=excel_row, column=COLUMN_MAP[key]).value = images[i] if i < len(images) else None

        written_ids.append(int(pid))
        excel_row += 1

    wb.save(out_path)
    print(f"  Wrote {len(written_ids)} products → {os.path.basename(out_path)}")
    return written_ids


# ── CLI entry point ────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Export products to Octopia .xlsm templates")
    parser.add_argument("--only-new", action="store_true", help="Only export products not yet exported (exported_at IS NULL)")
    args = parser.parse_args()

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print(f"Loading products (only_new={args.only_new})...")
    categories = load_products(only_new=args.only_new)

    if not categories:
        print("No products to export.")
        return

    all_written_ids = []
    for cat_id, cat_data in categories.items():
        cat_name  = cat_data["category_name"]
        products  = cat_data["products"]
        safe_name = make_safe_filename(str(cat_id), cat_name)
        out_path  = os.path.join(OUTPUT_DIR, f"{safe_name}.xlsm")

        print(f"\nCategory: {cat_id} — {cat_name} ({len(products)} products)")
        written = write_category_file(
            category_id=cat_id,
            category_name=cat_name,
            products=products,
            out_path=out_path,
            append_mode=args.only_new,
        )
        all_written_ids.extend(written)

    if all_written_ids:
        conn = sqlite3.connect(DB_PATH)
        now  = datetime.utcnow().isoformat()
        placeholders = ",".join("?" * len(all_written_ids))
        conn.execute(
            f"UPDATE product_fetched SET exported_at = ? WHERE product_id IN ({placeholders})",
            [now] + all_written_ids,
        )
        conn.commit()
        conn.close()
        print(f"\nMarked {len(all_written_ids)} product(s) as exported.")

    print(f"\nDone. Files in: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()COLUMN_MAP = {
    "sellerProductReference": 2,
    "title":                  3,
    "description":            4,
    "image_1":                5,
    "richMarketingDescription": 9,
    "image_2":               10,
    "image_3":               11,
    "image_4":               12,
    "image_5":               13,
    "image_6":               14,
}

REF_MAX         = 50
TITLE_MAX       = 132
DESCRIPTION_MAX = 2000
DATA_START_ROW  = 11   # rows 1-10 are template headers
DESC_MARKETING_MAX = 5000


# ── DB helpers ─────────────────────────────────────────────────────────────────

def load_products(db_path: str, only_new: bool = False) -> pd.DataFrame:
    """
    Load categorized products from SQLite.
    only_new=True  → only rows where exported_at IS NULL
    only_new=False → all categorized products
    """
    conn = sqlite3.connect(db_path)

    where = "WHERE ca.category_id IS NOT NULL"
    if only_new:
        where += " AND pf.exported_at IS NULL"

    query = f"""
        SELECT
            pf.product_id,
            pf.url,
            pf.title              AS original_title,
            pf.description        AS original_description,
            pf.images,
            pr.description_marketing,
            pr.enhanced_title,
            pr.enhanced_description,
            ca.assigned_category,
            ca.category_id,
            ca.llm_predicted_category,
            ca.similarity_score
        FROM product_fetched pf
        LEFT JOIN product_refined    pr ON pf.product_id = pr.product_id
        LEFT JOIN category_assignment ca ON pf.product_id = ca.product_id
        {where}
        ORDER BY ca.category_id
    """
    df = pd.read_sql_query(query, conn)
    conn.close()
    return df


def parse_images(images_raw) -> list:
    """
    Parse JSON images field and strip query params (?width=...&hash=...).
    Octopia rejects URLs with query strings.
    """
    if not images_raw:
        return []

    if isinstance(images_raw, list):
        imgs = images_raw
    else:
        try:
            imgs = json.loads(images_raw)
            if not isinstance(imgs, list):
                return []
        except Exception:
            return []

    return [url.split("?")[0].strip() for url in imgs if url]

def get_leaf_category(category_name: str) -> str:
    if not category_name:
        return ""
    return category_name.strip().split("/")[-1].strip()


def make_safe_filename(category_id: str, category_name: str) -> str:
    leaf = get_leaf_category(category_name)
    raw  = f"{category_id} - {leaf}" if leaf else category_id
    for ch in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
        raw = raw.replace(ch, '')
    return raw.strip()

# ── Template writer ────────────────────────────────────────────────────────────

def write_category_file(
    template_path: str,
    out_path: str,
    category_code: str,
    category_name: str,
    products: pd.DataFrame,
    append: bool = False,
) -> list:
    """
    Write products into an Octopia .xlsm template file.

    append=False → copy template fresh, write all products from row 11.
    append=True  → open existing file and append after the last used row.
                   Falls back to fresh copy if file does not exist.

    Returns list of product_ids written (used to mark exported_at in DB).
    """
    file_exists = os.path.exists(out_path)

    if append and file_exists:
        # Open existing file and find next empty row
        wb = load_workbook(out_path, keep_vba=True)
        ws = wb.active
        next_row = DATA_START_ROW
        for row in range(DATA_START_ROW, ws.max_row + 2):
            if all(ws.cell(row, c).value is None for c in [2, 3, 5]):
                next_row = row
                break
        print(f"  Appending from row {next_row} (file had {next_row - DATA_START_ROW} product(s) already)")
    else:
        # Fresh copy from template
        shutil.copy2(template_path, out_path)
        wb = load_workbook(out_path, keep_vba=True)
        ws = wb.active

        # Update Row 1 category header
        ws.cell(row=1, column=3).value = category_code
        ws.cell(row=1, column=4).value = category_name

        # Rename sheet (Excel max 31 chars, no special chars)
        safe_name = (category_name[:31]
                     .replace("/", "-").replace("\\", "-")
                     .replace("*", "").replace("?", "")
                     .replace(":", "").replace("[", "").replace("]", ""))
        ws.title = safe_name
        next_row = DATA_START_ROW

    # Write products
    written_ids = []
    for row_offset, (_, product) in enumerate(products.iterrows()):
        excel_row = next_row + row_offset

        title       = str(product.get("enhanced_title")       or product.get("original_title")       or "")
        description = str(product.get("enhanced_description") or product.get("original_description") or "")
        product_id  = str(product.get("product_id") or "")
        images      = parse_images(product.get("images"))
        desc_marketing = str(product.get("description_marketing") or "")[:DESC_MARKETING_MAX]
        #ws.cell(row=excel_row, column=COLUMN_MAP["descriptionMarketing"]).value = desc_marketing
        ws.cell(row=excel_row, column=COLUMN_MAP["richMarketingDescription"]).value = desc_marketing

        # Enforce Octopia character limits
        product_id  = product_id[:REF_MAX]
        title       = title[:TITLE_MAX].rstrip()
        description = description[:DESCRIPTION_MAX].rstrip()

        ws.cell(row=excel_row, column=COLUMN_MAP["sellerProductReference"]).value = product_id
        ws.cell(row=excel_row, column=COLUMN_MAP["title"]).value                  = title
        ws.cell(row=excel_row, column=COLUMN_MAP["description"]).value            = description

        img_keys = ["image_1", "image_2", "image_3", "image_4", "image_5", "image_6"]
        for i, key in enumerate(img_keys):
            if i < len(images):
                ws.cell(row=excel_row, column=COLUMN_MAP[key]).value = images[i]

        if len(images) > 6:
            print(f"    ⚠  product {product_id} has {len(images)} images — only first 6 written (template limit)")

        #written_ids.append(product.get("product_id"))
        written_ids.append(int(product.get("product_id")))


    wb.save(out_path)
    mode_label = "appended" if (append and file_exists) else "written"
    print(f"  ✅ {len(products)} product(s) {mode_label} → {out_path}")
    return written_ids


# ── Main (standalone CLI) ──────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Export products to Octopia template files")
    parser.add_argument("--template", default="data/pdt_template_fr-FR_20260305_090255.xlsm")
    parser.add_argument("--db",       default="data/products.db")
    parser.add_argument("--out",      default="data/output_templates")
    parser.add_argument("--only-new", action="store_true",
                        help="Only export products not yet exported (exported_at IS NULL)")
    args = parser.parse_args()

    if not os.path.exists(args.template):
        print(f"❌ Template not found: {args.template}")
        return
    if not os.path.exists(args.db):
        print(f"❌ Database not found: {args.db}")
        return

    os.makedirs(args.out, exist_ok=True)
    df = load_products(args.db, only_new=args.only_new)

    if df.empty:
        print("❌ No products to export.")
        return

    total_categories = df["category_id"].nunique()
    print(f"Found {len(df)} product(s) across {total_categories} category/ies\n")

    for category_id, group in df.groupby("category_id"):
        category_name = group.iloc[0]["assigned_category"] or str(category_id)
        #safe_filename = str(category_id).replace("/", "_").replace(" ", "_")
        safe_filename = make_safe_filename(str(category_id), category_name)
        out_path = os.path.join(args.out, f"{safe_filename}.xlsm")

        print(f"Category [{category_id}] {category_name} — {len(group)} product(s)")
        write_category_file(
            template_path=args.template,
            out_path=out_path,
            category_code=str(category_id),
            category_name=str(category_name),
            products=group,
            append=args.only_new,
        )

    print(f"\n✅ Done. {total_categories} file(s) written to '{args.out}/'")


if __name__ == "__main__":
    main()
